# -*- coding: utf-8 -*-
"""excel_deep_trace.py - deep formula tracing for produced Excel workbooks.

Library for the RWA Calculation webapp's ``/excel-trace/*`` routes.  Point it at a
workbook that carries **live Excel formulas** in its computed columns and it lets you:

* open any cell and see its formula broken down node-by-node, with the ``IF`` branch that
  is actually taken highlighted and every sub-value shown (``expression_tree``);
* walk an **entire row's** calculations at once (``row_journey``) - every formula cell in
  the row, not just one column;
* see what feeds a cell and what it feeds (``trace`` precedents / dependents);
* check which branch of every ``IF`` is live (``branch_analysis``) and what a ``VLOOKUP``
  matched (``lookup_analysis``);
* diff two cells (``compare_cells``);
* get workbook-wide lint findings (``findings``).

Reuses **webapp_core**: the formula-AST node classes from ``formula_translate`` and the
scalar helpers / value formatting from ``compute`` + ``formula_trace``.  No ``dataiku``
dependency - feed it a path or bytes.

Deploy: put ``webapp_core/`` and this file on the RWA project's Python libraries
(Project > Libraries).  The ``/excel-trace/*`` routes already call this class.
"""

from __future__ import annotations

import io
import math
import os
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter

from webapp_core.formula_translate import (
    Node, Num, Str, Bool, Nan, Paren, Unary, Bin, Cmp, ConcatOp, Call, If,
)
from webapp_core import compute
from webapp_core.formula_trace import (
    _repr, _dtype, _json, _truthy, _is_empty, _num, _is_nan, _values_close,
)

ENGINE_BUILD = "rwa-deep-trace 2026.09b"  # bump when the API surface changes
MAX_EXPANDED_RANGE_DEFAULT = 10000
MAX_TREE_DEPTH_DEFAULT = 8

_ERROR_LITERALS = {"#REF!", "#DIV/0!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#VALUE!",
                   "#SPILL!", "#CALC!", "#GETTING_DATA"}
_VOLATILE = {"NOW", "TODAY", "RAND", "RANDBETWEEN", "RANDARRAY", "OFFSET", "INDIRECT",
             "INFO", "CELL"}
_LOOKUP_FUNCS = {"VLOOKUP", "HLOOKUP", "XLOOKUP", "LOOKUP", "INDEX", "MATCH", "CHOOSE"}
_COND_FUNCS = {"IF", "IFS", "IFERROR", "IFNA", "SWITCH", "AND", "OR", "NOT"}
_AGG_FUNCS = {"SUM", "AVERAGE", "MIN", "MAX", "COUNT", "COUNTA", "PRODUCT", "MEDIAN"}


# =============================================================================
# A1 helpers
# =============================================================================
_CELL_RE = re.compile(r"^(\$?)([A-Za-z]{1,3})(\$?)(\d+)$")


def col_to_idx(letter: str) -> int:
    return column_index_from_string(letter.upper())


def idx_to_col(i: int) -> str:
    return get_column_letter(i)


def _unquote_sheet(s: str) -> str:
    s = s.strip()
    if s.startswith("'") and s.endswith("'"):
        s = s[1:-1].replace("''", "'")
    return s


def split_ref(ref: str) -> Tuple[Optional[str], str, int, bool, bool]:
    """``Sheet1!$B$5`` -> ``('Sheet1', 'B', 5, abs_col, abs_row)``."""
    ref = ref.strip()
    sheet = None
    if "!" in ref:
        sheet, ref = ref.rsplit("!", 1)
        sheet = _unquote_sheet(sheet)
    m = _CELL_RE.match(ref.strip())
    if not m:
        raise ValueError("bad cell reference %r" % ref)
    return sheet, m.group(2).upper(), int(m.group(4)), m.group(1) == "$", m.group(3) == "$"


# =============================================================================
# AST nodes specific to a real workbook (reuse everything else from webapp_core)
# =============================================================================
class CellRef(Node):
    def __init__(self, sheet, col, row, abs_col=False, abs_row=False):
        self.sheet = sheet
        self.col = col.upper()
        self.row = int(row)
        self.abs_col, self.abs_row = abs_col, abs_row

    @property
    def a1(self) -> str:
        return "%s%d" % (self.col, self.row)

    def key(self, default_sheet: Optional[str] = None) -> str:
        s = self.sheet or default_sheet
        return "%s!%s" % (s, self.a1) if s else self.a1

    def to_python(self) -> str:
        return "CELL(%r, %r)" % (self.sheet or "", self.a1)


class RangeRef(Node):
    def __init__(self, sheet, c1, r1, c2, r2):
        self.sheet = sheet
        self.c1, self.c2 = c1.upper(), c2.upper()
        self.r1, self.r2 = int(r1), int(r2)

    @property
    def whole_column(self) -> bool:
        return self.r2 < 0

    @property
    def a1(self) -> str:
        if self.whole_column:
            return "%s:%s" % (self.c1, self.c2)
        return "%s%d:%s%d" % (self.c1, self.r1, self.c2, self.r2)

    def to_python(self) -> str:
        return "RANGE(%r, %r)" % (self.sheet or "", self.a1)


class NameRef(Node):
    def __init__(self, name):
        self.name = name

    def to_python(self) -> str:
        return "NAME(%r)" % self.name


class ErrorLit(Node):
    def __init__(self, text):
        self.text = text

    def to_python(self) -> str:
        return self.text


# =============================================================================
# Excel formula lexer + parser  ->  the shared AST
# =============================================================================
_XL_SPEC = [
    ("WS", r"[ \t\r\n]+"),
    ("STRING", r'"(?:[^"]|"")*"'),
    ("ERROR", r"#REF!|#DIV/0!|#N/A|#NAME\?|#NULL!|#NUM!|#VALUE!|#SPILL!|#CALC!|#GETTING_DATA"),
    ("SHEETRANGE", r"(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!\$?[A-Za-z]{1,3}\$?\d+:\$?[A-Za-z]{1,3}\$?\d+"),
    ("SHEETCOLRANGE", r"(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!\$?[A-Za-z]{1,3}:\$?[A-Za-z]{1,3}"),
    ("SHEETCELL", r"(?:'(?:[^']|'')+'|[A-Za-z_][A-Za-z0-9_.]*)!\$?[A-Za-z]{1,3}\$?\d+"),
    ("RANGE", r"\$?[A-Za-z]{1,3}\$?\d+:\$?[A-Za-z]{1,3}\$?\d+"),
    ("COLRANGE", r"\$?[A-Za-z]{1,3}:\$?[A-Za-z]{1,3}"),
    ("NUMBER", r"\d+(?:\.\d+)?(?:[eE][+-]?\d+)?|\.\d+"),
    ("FUNC", r"[A-Za-z_][A-Za-z0-9_.]*(?=\s*\()"),
    ("CELLREF", r"\$?[A-Za-z]{1,3}\$?\d+"),
    ("IDENT", r"[A-Za-z_][A-Za-z0-9_.]*"),
    ("OP", r"<>|<=|>=|[-+*/^&=<>%]"),
    ("LP", r"\("),
    ("RP", r"\)"),
    ("COMMA", r"[,;]"),
]
_XL_RE = re.compile("|".join("(?P<%s>%s)" % (n, p) for n, p in _XL_SPEC))


class _Tok:
    __slots__ = ("kind", "val")

    def __init__(self, kind, val):
        self.kind, self.val = kind, val

    def __repr__(self):  # pragma: no cover
        return "%s:%s" % (self.kind, self.val)


def _xl_tokenize(s: str) -> List[_Tok]:
    out, pos = [], 0
    while pos < len(s):
        m = _XL_RE.match(s, pos)
        if not m:
            raise ValueError("cannot tokenize near %r" % s[pos:pos + 16])
        pos = m.end()
        if m.lastgroup != "WS":
            out.append(_Tok(m.lastgroup, m.group()))
    return out


class _XlParser:
    def __init__(self, tokens: List[_Tok], default_sheet: Optional[str]):
        self.toks = tokens
        self.i = 0
        self.default_sheet = default_sheet
        self.notes: List[str] = []

    def _peek(self):
        return self.toks[self.i] if self.i < len(self.toks) else None

    def _next(self):
        t = self.toks[self.i]
        self.i += 1
        return t

    def _eat(self, kind):
        t = self._peek()
        if not t or t.kind != kind:
            raise ValueError("expected %s, got %s" % (kind, t))
        return self._next()

    def _op(self, *vals):
        t = self._peek()
        return bool(t and t.kind == "OP" and t.val in vals)

    def parse(self) -> Node:
        node = self._concat()
        if self.i != len(self.toks):
            self.notes.append("trailing tokens ignored")
        return node

    def _concat(self):
        left = self._compare()
        while self._op("&"):
            self._next()
            left = ConcatOp(left, self._compare())
        return left

    def _compare(self):
        left = self._addsub()
        while self._op("=", "<>", "<", ">", "<=", ">="):
            op = self._next().val
            py = {"=": "==", "<>": "!="}.get(op, op)
            left = Cmp(py, left, self._addsub())
        return left

    def _addsub(self):
        left = self._muldiv()
        while self._op("+", "-"):
            op = self._next().val
            left = Bin(op, left, self._muldiv())
        return left

    def _muldiv(self):
        left = self._power()
        while self._op("*", "/"):
            op = self._next().val
            left = Bin(op, left, self._power())
        return left

    def _power(self):
        left = self._postfix()
        while self._op("^"):
            self._next()
            left = Bin("**", left, self._postfix())
        return left

    def _postfix(self):
        node = self._unary()
        while self._op("%"):
            self._next()
            node = Bin("/", node, Num("100"))
        return node

    def _unary(self):
        if self._op("-", "+"):
            op = self._next().val
            return Unary(op, self._unary())
        return self._primary()

    def _primary(self):
        t = self._peek()
        if t is None:
            raise ValueError("unexpected end of formula")
        if t.kind == "LP":
            self._next()
            inner = self._concat()
            self._eat("RP")
            return Paren(inner)
        if t.kind == "NUMBER":
            return Num(self._next().val)
        if t.kind == "STRING":
            return Str(self._next().val[1:-1].replace('""', '"'))
        if t.kind == "ERROR":
            return ErrorLit(self._next().val)
        if t.kind == "FUNC":
            return self._func()
        if t.kind in ("SHEETRANGE", "RANGE"):
            return self._range(self._next().val)
        if t.kind in ("SHEETCOLRANGE", "COLRANGE"):
            return self._colrange(self._next().val)
        if t.kind in ("SHEETCELL", "CELLREF"):
            return self._cell(self._next().val)
        if t.kind == "IDENT":
            name = self._next().val
            if name.upper() in ("TRUE", "FALSE"):
                return Bool(name.capitalize())
            return NameRef(name)
        raise ValueError("unexpected token %s" % t)

    def _cell(self, raw):
        sheet, col, row, ac, ar = split_ref(raw)
        return CellRef(sheet or self.default_sheet, col, row, ac, ar)

    def _range(self, raw):
        sheet, body = (raw.rsplit("!", 1) + [None])[:2] if "!" in raw else (None, raw)
        if sheet is not None:
            sheet = _unquote_sheet(sheet)
        a, b = body.split(":")
        _, c1, r1, _, _ = split_ref(a)
        _, c2, r2, _, _ = split_ref(b)
        return RangeRef(sheet or self.default_sheet, c1, r1, c2, r2)

    def _colrange(self, raw):
        """A whole-column range like ``Map!$A:$C`` — rows resolved lazily (r1=1, r2=-1)."""
        if "!" in raw:
            sheet, body = raw.rsplit("!", 1)
            sheet = _unquote_sheet(sheet)
        else:
            sheet, body = None, raw
        a, b = body.replace("$", "").split(":")
        return RangeRef(sheet or self.default_sheet, a, 1, b, -1)

    def _func(self):
        fname = self._next().val.upper()
        self._eat("LP")
        args: List[Node] = []
        if self._peek() and self._peek().kind != "RP":
            while True:
                args.append(self._concat())
                if self._peek() and self._peek().kind == "COMMA":
                    self._next()
                    continue
                break
        self._eat("RP")

        if fname == "IF":
            if len(args) >= 3:
                return If(args[0], args[1], args[2])
            if len(args) == 2:
                return If(args[0], args[1], None)
            return Call("IF", args)
        if fname == "IFS":
            node: Node = ErrorLit("#N/A")
            for k in range(len(args) - 2, -1, -2):
                node = If(args[k], args[k + 1], node)
            return node
        if fname == "SWITCH" and len(args) >= 3:
            subject = args[0]
            default = args[-1] if len(args) % 2 == 0 else ErrorLit("#N/A")
            pairs = args[1:-1] if len(args) % 2 == 0 else args[1:]
            node = default
            for k in range(len(pairs) - 2, -1, -2):
                node = If(Cmp("==", subject, pairs[k]), pairs[k + 1], node)
            return node
        return Call(fname, args)


def parse_excel(formula: str, default_sheet: Optional[str] = None) -> Tuple[Node, List[str]]:
    """Parse a real Excel formula string into the shared AST. Raises on failure."""
    s = formula.strip()
    if s.startswith("="):
        s = s[1:]
    if not s:
        raise ValueError("empty formula")
    p = _XlParser(_xl_tokenize(s), default_sheet)
    return p.parse(), p.notes


# =============================================================================
# runtime error sentinel
# =============================================================================
class XlError:
    __slots__ = ("code",)

    def __init__(self, code):
        self.code = code

    def __bool__(self):
        return False

    def __repr__(self):
        return self.code

    def __eq__(self, other):
        return isinstance(other, XlError) and other.code == self.code

    def __hash__(self):
        return hash(self.code)


def _is_err(v) -> bool:
    return isinstance(v, XlError)


def _f(x) -> float:
    if x is None or x == "":
        return 0.0
    if isinstance(x, bool):
        return 1.0 if x else 0.0
    if isinstance(x, (int, float)):
        return float(x)
    return float(str(x).replace(",", "").strip())


def _loose_eq(a, b) -> bool:
    if _is_empty(a) and _is_empty(b):
        return True
    an, bn = _num(a), _num(b)
    if isinstance(an, (int, float)) and isinstance(bn, (int, float)) \
            and not isinstance(a, bool) and not isinstance(b, bool):
        return abs(float(an) - float(bn)) < 1e-9
    return str(a).strip().lower() == str(b).strip().lower()


def _xl_cmp(op, a, b) -> Any:
    if _is_err(a):
        return a
    if _is_err(b):
        return b
    an, bn = _num(a), _num(b)
    numeric = (isinstance(an, (int, float)) and isinstance(bn, (int, float))
               and not isinstance(a, bool) and not isinstance(b, bool))
    if _is_empty(a) and isinstance(bn, (int, float)):
        an, numeric = 0, True
    if _is_empty(b) and isinstance(an, (int, float)):
        bn, numeric = 0, True
    if numeric:
        x, y = float(an), float(bn)
    else:
        x, y = str("" if a is None else a).strip().lower(), str("" if b is None else b).strip().lower()
    return {"==": x == y, "!=": x != y, "<": x < y, ">": x > y, "<=": x <= y, ">=": x >= y}[op]


# =============================================================================
# Excel-ish source rendering of a node
# =============================================================================
_BIN_SRC = {"**": "^"}
_CMP_SRC = {"==": "=", "!=": "<>"}


def src(n: Node) -> str:
    if isinstance(n, Num):
        return n.text
    if isinstance(n, Str):
        return '"%s"' % n.value
    if isinstance(n, Bool):
        return n.py.upper()
    if isinstance(n, Nan):
        return ""
    if isinstance(n, ErrorLit):
        return n.text
    if isinstance(n, NameRef):
        return n.name
    if isinstance(n, CellRef):
        return ("%s!" % n.sheet if n.sheet else "") + n.a1
    if isinstance(n, RangeRef):
        return ("%s!" % n.sheet if n.sheet else "") + n.a1
    if isinstance(n, Paren):
        return "(%s)" % src(n.inner)
    if isinstance(n, Unary):
        return "%s%s" % (n.op, src(n.operand))
    if isinstance(n, Bin):
        return "%s %s %s" % (src(n.left), _BIN_SRC.get(n.op, n.op), src(n.right))
    if isinstance(n, Cmp):
        return "%s %s %s" % (src(n.left), _CMP_SRC.get(n.op, n.op), src(n.right))
    if isinstance(n, ConcatOp):
        return "%s & %s" % (src(n.left), src(n.right))
    if isinstance(n, If):
        parts = [src(n.cond), src(n.then)] + ([src(n.els)] if n.els is not None else [])
        return "IF(%s)" % ", ".join(parts)
    if isinstance(n, Call):
        return "%s(%s)" % (n.name, ", ".join(src(a) for a in n.args))
    return "?"


# =============================================================================
# scalar helpers reused / adapted
# =============================================================================
def _agg_values(items: List[Any]) -> List[float]:
    out = []
    for v in items:
        if _is_err(v) or _is_empty(v) or isinstance(v, bool):
            continue
        n = _num(v)
        if isinstance(n, (int, float)):
            out.append(float(n))
    return out


_SCALAR = {
    "ROUND": lambda x, d=0: round(_f(x), int(_f(d))),
    "ROUNDUP": lambda x, d=0: math.ceil(_f(x) * 10 ** int(_f(d))) / 10 ** int(_f(d)),
    "ROUNDDOWN": lambda x, d=0: math.floor(_f(x) * 10 ** int(_f(d))) / 10 ** int(_f(d)),
    "INT": lambda x: math.floor(_f(x)),
    "ABS": lambda x: abs(_f(x)),
    "MOD": lambda a, b: _f(a) % _f(b) if _f(b) != 0 else XlError("#DIV/0!"),
    "POWER": lambda a, b: _f(a) ** _f(b),
    "SQRT": lambda x: math.sqrt(_f(x)) if _f(x) >= 0 else XlError("#NUM!"),
    "CEILING": lambda x, s=1: math.ceil(_f(x) / _f(s)) * _f(s) if _f(s) else XlError("#DIV/0!"),
    "FLOOR": lambda x, s=1: math.floor(_f(x) / _f(s)) * _f(s) if _f(s) else XlError("#DIV/0!"),
    "LEN": lambda x: len("" if x is None else str(x)),
    "LEFT": lambda x, n=1: ("" if x is None else str(x))[: int(_f(n))],
    "RIGHT": lambda x, n=1: ("" if x is None else str(x))[-int(_f(n)):] if int(_f(n)) else "",
    "MID": lambda x, s, n: ("" if x is None else str(x))[int(_f(s)) - 1: int(_f(s)) - 1 + int(_f(n))],
    "UPPER": lambda x: ("" if x is None else str(x)).upper(),
    "LOWER": lambda x: ("" if x is None else str(x)).lower(),
    "TRIM": lambda x: re.sub(r"\s+", " ", ("" if x is None else str(x))).strip(),
    "TEXT": lambda x, *_: "" if x is None else str(x),
    "VALUE": lambda x: _f(x),
    "CONCATENATE": lambda *a: "".join("" if v is None else str(v) for v in a),
    "CONCAT": lambda *a: "".join("" if v is None else str(v) for v in a),
    "SUBSTITUTE": lambda s, o, nw, *_: ("" if s is None else str(s)).replace(str(o), str(nw)),
    "ISBLANK": lambda x: _is_empty(x),
    "ISNUMBER": lambda x: isinstance(_num(x), (int, float)) and not isinstance(x, bool) and not _is_empty(x),
    "ISTEXT": lambda x: isinstance(x, str) and x != "",
    "ISERROR": lambda x: _is_err(x),
    "ISERR": lambda x: _is_err(x) and x.code != "#N/A",
    "ISNA": lambda x: _is_err(x) and x.code == "#N/A",
    "N": lambda x: _f(x) if not _is_empty(x) else 0.0,
    "SIGN": lambda x: (_f(x) > 0) - (_f(x) < 0),
    "TRUE": lambda: True,
    "FALSE": lambda: False,
}
_SCALAR_MIN_ARGS = {"MID": 3}


# =============================================================================
# records + findings
# =============================================================================
@dataclass
class Record:
    key: str
    sheet: str
    cell: str
    coordinate: str
    formula: str
    cached_value: Any
    functions: List[str]
    family: str
    formula_length: int
    dependency_count: int
    is_conditional: bool
    is_lookup: bool
    is_volatile: bool
    parsed: bool = True
    precedents: set = field(default_factory=set)
    ranges: list = field(default_factory=list)
    root: Any = field(default=None, repr=False)


@dataclass
class Finding:
    severity: str
    code: str
    cell: str
    message: str
    evidence: Any = None


_FUNC_CALL_RE = re.compile(r"([A-Za-z_][A-Za-z0-9_.]*)\s*\(")


def _functions_in(formula: str) -> set:
    return {m.group(1).upper() for m in _FUNC_CALL_RE.finditer(formula)}


def _family(funcs: set, root: Optional[Node]) -> str:
    if funcs & _LOOKUP_FUNCS:
        return "lookup"
    if funcs & {"IF", "IFS", "IFERROR", "IFNA", "SWITCH"}:
        return "conditional"
    if funcs & (_AGG_FUNCS | {"SUMIF", "SUMIFS", "COUNTIF", "COUNTIFS", "SUMPRODUCT", "AGGREGATE"}):
        return "aggregation"
    if funcs & {"CONCAT", "CONCATENATE", "TEXT", "LEFT", "RIGHT", "MID", "TRIM", "UPPER",
                "LOWER", "SUBSTITUTE"}:
        return "text"
    if funcs <= {"ROUND", "ROUNDUP", "ROUNDDOWN", "INT", "ABS", "MOD", "POWER", "SQRT",
                 "CEILING", "FLOOR", "SIGN", "N", "VALUE"}:
        # only arithmetic-shaping helpers (or none) -> it's an arithmetic formula
        if root is not None and any(isinstance(x, (Bin, Unary)) for x in root.walk()):
            return "arithmetic"
        if not funcs and root is not None:
            return "arithmetic" if isinstance(root, (Bin, Unary, Paren)) else "reference"
        if funcs:
            return "arithmetic"
    if not funcs and root is not None:
        return "arithmetic" if isinstance(root, (Bin, Unary, Paren)) else "reference"
    return "other"


def _deps(root: Node, default_sheet: str) -> Tuple[set, list]:
    cells, ranges = set(), []
    for n in root.walk():
        if isinstance(n, CellRef):
            cells.add("%s!%s" % (n.sheet or default_sheet, n.a1))
        elif isinstance(n, RangeRef):
            ranges.append((n.sheet or default_sheet, n.c1, n.r1, n.c2, n.r2))
    return cells, ranges


def _range_keys(sheet, c1, r1, c2, r2, cap=400):
    lo_c, hi_c = sorted((col_to_idx(c1), col_to_idx(c2)))
    if int(r2) < 1:  # whole-column range - just link the header cells
        return {"%s!%s1" % (sheet, idx_to_col(c)) for c in range(lo_c, hi_c + 1)}
    lo_r, hi_r = sorted((int(r1), int(r2)))
    if (hi_c - lo_c + 1) * (hi_r - lo_r + 1) > cap:
        # sample the corners so the graph still links, without exploding
        pts = {(lo_c, lo_r), (hi_c, lo_r), (lo_c, hi_r), (hi_c, hi_r)}
        return {"%s!%s%d" % (sheet, idx_to_col(c), r) for c, r in pts}
    return {"%s!%s%d" % (sheet, idx_to_col(c), r)
            for c in range(lo_c, hi_c + 1) for r in range(lo_r, hi_r + 1)}


# =============================================================================
# the tracing evaluator
# =============================================================================
class _Eval:
    def __init__(self, engine: "ExcelDeepTraceEngine", trace: bool = False,
                 max_depth: int = MAX_TREE_DEPTH_DEFAULT, if_ids: Optional[dict] = None):
        self.e = engine
        self.trace = trace
        self.max_depth = max_depth
        self.if_ids = if_ids or {}
        self.notes: List[str] = []
        self._nid = 0
        self._stack: set = set()

    # ---- node factory ----------------------------------------------------
    def _mk(self, kind, node, value, role, children=None, evaluated=True, **detail):
        self._nid += 1
        return {
            "id": self._nid,
            "kind": kind,
            "role": role,
            "excel": src(node),
            "value": _json(value),
            "value_repr": (value.code if _is_err(value) else _repr(value)),
            "dtype": ("error" if _is_err(value) else _dtype(value)),
            "evaluated": evaluated,
            "children": children or [],
            "detail": {k: v for k, v in detail.items() if v is not None},
        }

    def _skeleton(self, node, role):
        self._nid += 1
        d = {
            "id": self._nid, "kind": "skipped", "role": role, "excel": src(node),
            "value": None, "value_repr": "not evaluated", "dtype": "skipped",
            "evaluated": False, "children": [self._skeleton(c, "arg") for c in node.kids()],
            "detail": {},
        }
        if isinstance(node, If) and id(node) in self.if_ids:
            d["detail"]["branch_id"] = self.if_ids[id(node)]
        return d

    # ---- pure value (no tree) -----------------------------------------
    def value(self, node) -> Any:
        _, v = self.ev(node, 0, "root")
        return v

    # ---- main dispatch ----------------------------------------------------
    def ev(self, n: Node, depth: int, role: str = "root") -> Tuple[Optional[dict], Any]:
        if isinstance(n, Paren):
            return self.ev(n.inner, depth, role)
        if isinstance(n, Num):
            v = float(n.text) if ("." in n.text or "e" in n.text.lower()) else int(n.text)
            return self._mk("number", n, v, role), v
        if isinstance(n, Str):
            return self._mk("text", n, n.value, role), n.value
        if isinstance(n, Bool):
            v = n.py == "True"
            return self._mk("bool", n, v, role), v
        if isinstance(n, Nan):
            return self._mk("empty", n, None, role), None
        if isinstance(n, ErrorLit):
            v = XlError(n.text)
            return self._mk("error", n, v, role, error=n.text), v
        if isinstance(n, NameRef):
            v = self.e._defined_name(n.name)
            if v is None:
                self.notes.append("name %r is not a defined name in this workbook" % n.name)
                v = XlError("#NAME?")
            return self._mk("name", n, v, role, name=n.name), v
        if isinstance(n, CellRef):
            return self._cell(n, depth, role)
        if isinstance(n, RangeRef):
            return self._mk("range", n, "%s cells" % self._range_size(n), role,
                            a1=n.a1, sheet=n.sheet), self.e._grid_or_err(n)
        if isinstance(n, Unary):
            cn, cv = self.ev(n.operand, depth, "operand")
            v = cv if _is_err(cv) else (-_f(cv) if n.op == "-" else _f(cv))
            return self._mk("unary", n, v, role, [cn], operator=n.op), v
        if isinstance(n, Bin):
            return self._bin(n, depth, role)
        if isinstance(n, Cmp):
            cn, cv = self.ev(n.left, depth, "left")
            dn, dv = self.ev(n.right, depth, "right")
            v = _xl_cmp(n.op, cv, dv)
            return self._mk("compare", n, v, role, [cn, dn], operator=n.op), v
        if isinstance(n, ConcatOp):
            cn, cv = self.ev(n.left, depth, "left")
            dn, dv = self.ev(n.right, depth, "right")
            if _is_err(cv):
                return self._mk("concat", n, cv, role, [cn, dn]), cv
            if _is_err(dv):
                return self._mk("concat", n, dv, role, [cn, dn]), dv
            v = ("" if cv is None else str(cv)) + ("" if dv is None else str(dv))
            return self._mk("concat", n, v, role, [cn, dn]), v
        if isinstance(n, If):
            return self._if(n, depth, role)
        if isinstance(n, Call):
            return self._call(n, depth, role)
        return self._mk("unsupported", n, XlError("#N/A"), role), XlError("#N/A")

    # ---- leaves --------------------------------------------------------
    def _cell(self, n: CellRef, depth: int, role: str):
        sheet = n.sheet or self.e._book.default_sheet
        key = "%s!%s" % (sheet, n.a1)
        rec = self.e.records.get(key)
        raw = self.e._book.cached(sheet, n.a1)
        detail = {"key": key, "cell": n.a1, "sheet": sheet, "is_formula": rec is not None}
        if isinstance(raw, str) and raw in _ERROR_LITERALS:
            raw = XlError(raw)
        # Recurse into a referenced formula cell whenever we have no cached value for it
        # (formula-only workbooks) or we're building the visual tree. Non-trace mode skips
        # the recursion only when the cell already has a usable cached value.
        recurse = (rec is not None and rec.root is not None and depth < self.max_depth
                   and (self.trace or _is_empty(raw)))
        if recurse:
            if key in self._stack:
                return self._mk("cell", n, XlError("#REF!"), role,
                                error="circular reference", **detail), XlError("#REF!")
            self._stack.add(key)
            try:
                sub, subval = self.ev(rec.root, depth + 1, "ref")
            finally:
                self._stack.discard(key)
            node = self._mk("cell", n, subval, role, [sub] if self.trace else [], **detail)
            return node, subval
        if rec is not None:
            detail["collapsed_formula"] = rec.formula
        node = self._mk("cell", n, raw, role, [], **detail)
        return node, raw

    def _range_size(self, n: RangeRef) -> int:
        lo_c, hi_c = sorted((col_to_idx(n.c1), col_to_idx(n.c2)))
        if n.whole_column:
            try:
                grid = self.e._grid(n, self.e.max_expanded_range)
                return sum(len(r) for r in grid)
            except Exception:  # noqa: BLE001
                return (hi_c - lo_c + 1)
        lo_r, hi_r = sorted((n.r1, n.r2))
        return (hi_c - lo_c + 1) * (hi_r - lo_r + 1)

    # ---- operators ---------------------------------------------------
    def _bin(self, n: Bin, depth: int, role: str):
        ln, lv = self.ev(n.left, depth, "left")
        rn, rv = self.ev(n.right, depth, "right")
        if _is_err(lv):
            return self._mk("op", n, lv, role, [ln, rn], operator=n.op), lv
        if _is_err(rv):
            return self._mk("op", n, rv, role, [ln, rn], operator=n.op), rv
        try:
            a, b = _f(lv), _f(rv)
            if n.op == "+":
                v = a + b
            elif n.op == "-":
                v = a - b
            elif n.op == "*":
                v = a * b
            elif n.op == "/":
                v = XlError("#DIV/0!") if b == 0 else a / b
            elif n.op == "**":
                v = a ** b
            else:  # pragma: no cover
                v = XlError("#VALUE!")
        except (ValueError, TypeError):
            v = XlError("#VALUE!")
        return self._mk("op", n, v, role, [ln, rn], operator=n.op), v

    # ---- branches --------------------------------------------------
    def _if(self, n: If, depth: int, role: str):
        cn, cv = self.ev(n.cond, depth, "cond")
        if _is_err(cv):
            return self._mk("if", n, cv, role, [cn], branch="error",
                            branch_id=self.if_ids.get(id(n))), cv
        taken_then = _truthy(cv)
        kids = [cn]
        if taken_then:
            tn, val = self.ev(n.then, depth, "then")
            tn["taken"] = True
            kids.append(tn)
            if n.els is not None:
                sk = self._skeleton(n.els, "else")
                sk["taken"] = False
                kids.append(sk)
        else:
            sk = self._skeleton(n.then, "then")
            sk["taken"] = False
            kids.append(sk)
            if n.els is not None:
                en, val = self.ev(n.els, depth, "else")
                en["taken"] = True
                kids.append(en)
            else:
                val = False
        node = self._mk("if", n, val, role, kids,
                        branch=("true" if taken_then else "false"),
                        branch_id=self.if_ids.get(id(n)),
                        cond_empty=_is_empty(cv) or None)
        return node, val

    # ---- calls ----------------------------------------------------
    def _call(self, n: Call, depth: int, role: str):
        name = n.name.upper()

        if name in ("IFERROR", "IFNA"):
            an, av = self.ev(n.args[0], depth, "value")
            bad = _is_err(av) if name == "IFERROR" else (_is_err(av) and av.code == "#N/A")
            if bad:
                bn, bv = self.ev(n.args[1], depth, "fallback")
                an["taken"] = False
                bn["taken"] = True
                return self._mk("iferror", n, bv, role, [an, bn],
                                caught=(av.code if _is_err(av) else None)), bv
            sk = self._skeleton(n.args[1], "fallback")
            an["taken"] = True
            sk["taken"] = False
            return self._mk("iferror", n, av, role, [an, sk]), av

        if name in ("AND", "OR"):
            want = name == "OR"
            kids, res, sc = [], (not want), None
            for i, a in enumerate(n.args):
                an, av = self.ev(a, depth, "arg")
                kids.append(an)
                if _is_err(av):
                    return self._mk("logical", n, av, role, kids, operator=name), av
                if _truthy(av) == want:
                    res, sc = want, i
                    for rest in n.args[i + 1:]:
                        s = self._skeleton(rest, "arg")
                        s["skipped"] = True
                        kids.append(s)
                    break
            return self._mk("logical", n, res, role, kids, operator=name,
                            short_circuit_at=sc), res

        if name == "NOT":
            an, av = self.ev(n.args[0], depth, "arg")
            v = av if _is_err(av) else (not _truthy(av))
            return self._mk("logical", n, v, role, [an], operator="NOT"), v

        if name == "VLOOKUP":
            return self._vlookup(n, depth, role)
        if name in ("INDEX", "MATCH"):
            return self._index_match(n, depth, role)

        if name in _AGG_FUNCS or name == "SUMPRODUCT":
            return self._agg(n, depth, role, name)

        # generic scalar helper
        fn = _SCALAR.get(name)
        kids, vals = [], []
        for a in n.args:
            an, av = self.ev(a, depth, "arg")
            kids.append(an)
            vals.append(av)
        propagate = name not in ("ISERROR", "ISERR", "ISNA", "ISBLANK", "IFERROR", "IFNA", "N")
        if propagate:
            for v in vals:
                if _is_err(v):
                    return self._mk("func", n, v, role, kids, func=name), v
        if fn is None:
            self.notes.append("%s() is not supported by the tracer - Excel's cached "
                              "value is used at the parent cell" % name)
            v = XlError("#N/A")
            return self._mk("func", n, v, role, kids, func=name, unsupported=True), v
        try:
            v = fn(*vals) if vals else fn()
        except Exception as exc:  # noqa: BLE001
            self.notes.append("%s(): %s" % (name, exc))
            v = XlError("#VALUE!")
        return self._mk("func", n, v, role, kids, func=name), v

    # ---- lookups -------------------------------------------------
    def _vlookup(self, n: Call, depth: int, role: str):
        kn, kv = self.ev(n.args[0], depth, "key")
        table = n.args[1] if len(n.args) > 1 else None
        col_i = int(_f(self.ev(n.args[2], depth, "col")[1])) if len(n.args) > 2 else 2
        approx = _truthy(self.ev(n.args[3], depth, "range_lookup")[1]) if len(n.args) > 3 else True
        detail = {"function": "VLOOKUP", "lookup_value": _json(kv), "lookup_repr": _repr(kv),
                  "col_index": col_i, "approximate": bool(approx),
                  "table_range": src(table) if table is not None else None}
        if not isinstance(table, RangeRef):
            return self._mk("lookup", n, XlError("#REF!"), role, [kn],
                            error="lookup table is not a range", **detail), XlError("#REF!")
        try:
            grid = self.e._grid(table, self.e.max_expanded_range)
        except Exception as exc:  # noqa: BLE001
            detail["error"] = str(exc)
            return self._mk("lookup", n, XlError("#REF!"), role, [kn], **detail), XlError("#REF!")
        keys = [row[0] if row else None for row in grid]
        match = None
        if approx:
            for i, k in enumerate(keys):
                try:
                    if _f(k) <= _f(kv):
                        match = i
                except (ValueError, TypeError):
                    continue
        else:
            for i, k in enumerate(keys):
                if _loose_eq(k, kv):
                    match = i
                    break
        if match is None:
            v = XlError("#N/A")
            detail["match_found"] = False
        else:
            row = grid[match]
            v = row[col_i - 1] if 0 < col_i <= len(row) else XlError("#REF!")
            if isinstance(v, str) and v in _ERROR_LITERALS:
                v = XlError(v)
            detail.update({"match_found": True, "matched_row_index": match,
                           "matched_key": _json(row[0] if row else None),
                           "matched_row": [_json(x) for x in row],
                           "returned_value": _json(v)})
        detail["table_preview"] = [
            {"cells": [_json(x) for x in r], "matched": i == match}
            for i, r in enumerate(grid[:40])
        ]
        return self._mk("lookup", n, v, role, [kn], **detail), v

    def _index_match(self, n: Call, depth: int, role: str):
        name = n.name.upper()
        kids = []
        for a in n.args:
            an, _ = self.ev(a, depth, "arg")
            kids.append(an)
        try:
            if name == "MATCH" and len(n.args) >= 2 and isinstance(n.args[1], RangeRef):
                kv = self.ev(n.args[0], depth, "key")[1]
                grid = self.e._grid(n.args[1], self.e.max_expanded_range)
                flat = [c for r in grid for c in r]
                mtype = int(_f(self.ev(n.args[2], depth, "type")[1])) if len(n.args) > 2 else 1
                idx = None
                if mtype == 0:
                    for i, c in enumerate(flat):
                        if _loose_eq(c, kv):
                            idx = i + 1
                            break
                else:
                    for i, c in enumerate(flat):
                        try:
                            if (mtype > 0 and _f(c) <= _f(kv)) or (mtype < 0 and _f(c) >= _f(kv)):
                                idx = i + 1
                        except (ValueError, TypeError):
                            continue
                v = idx if idx else XlError("#N/A")
                return self._mk("lookup", n, v, role, kids, function="MATCH",
                                match_index=idx), v
            if name == "INDEX" and n.args and isinstance(n.args[0], RangeRef):
                grid = self.e._grid(n.args[0], self.e.max_expanded_range)
                r = int(_f(self.ev(n.args[1], depth, "row")[1])) if len(n.args) > 1 else 1
                c = int(_f(self.ev(n.args[2], depth, "col")[1])) if len(n.args) > 2 else 1
                if len(grid) == 1:
                    r, c = 1, max(r, c)
                elif len(grid[0]) == 1:
                    c = 1
                v = grid[r - 1][c - 1] if 0 < r <= len(grid) and 0 < c <= len(grid[r - 1]) else XlError("#REF!")
                return self._mk("lookup", n, v, role, kids, function="INDEX"), v
        except Exception as exc:  # noqa: BLE001
            self.notes.append("%s(): %s" % (name, exc))
        v = XlError("#N/A")
        self.notes.append("%s() could not be resolved by the tracer" % name)
        return self._mk("lookup", n, v, role, kids, function=name, unsupported=True), v

    # ---- aggregation ---------------------------------------------
    def _agg(self, n: Call, depth: int, role: str, name: str):
        kids, flat = [], []
        for a in n.args:
            if isinstance(a, RangeRef):
                self._nid += 1
                kids.append({"id": self._nid, "kind": "range", "role": "arg",
                             "excel": src(a), "value": None,
                             "value_repr": "%d cells" % self._range_size(a),
                             "dtype": "range", "evaluated": True, "children": [],
                             "detail": {"a1": a.a1}})
                try:
                    grid = self.e._grid(a, self.e.max_expanded_range)
                    flat.extend(c for row in grid for c in row)
                except Exception as exc:  # noqa: BLE001
                    self.notes.append("%s: %s" % (name, exc))
            else:
                an, av = self.ev(a, depth, "arg")
                kids.append(an)
                flat.append(av)
        nums = _agg_values(flat)
        if name == "SUM" or name == "SUMPRODUCT":
            v = sum(nums)
        elif name == "PRODUCT":
            v = math.prod(nums) if nums else 0
        elif name == "AVERAGE":
            v = sum(nums) / len(nums) if nums else XlError("#DIV/0!")
        elif name == "MIN":
            v = min(nums) if nums else 0
        elif name == "MAX":
            v = max(nums) if nums else 0
        elif name == "MEDIAN":
            s = sorted(nums)
            v = (s[len(s) // 2] if len(s) % 2 else (s[len(s) // 2 - 1] + s[len(s) // 2]) / 2) if s else XlError("#NUM!")
        elif name == "COUNT":
            v = len(nums)
        elif name == "COUNTA":
            v = sum(1 for x in flat if not _is_empty(x))
        else:  # pragma: no cover
            v = XlError("#N/A")
        return self._mk("agg", n, v, role, kids, func=name, n_values=len(nums)), v


# =============================================================================
# narration
# =============================================================================
def _narrate(root: dict, key: str, cached: Any) -> List[str]:
    out = ["%s = %s." % (key, root.get("value_repr"))]

    def walk(nd):
        k = nd.get("kind")
        if k == "if":
            cond = nd["children"][0]
            br = nd["detail"].get("branch", "?")
            extra = " (empty, treated as TRUE)" if nd["detail"].get("cond_empty") else ""
            out.append("Because %s is %s%s, the %s branch is used."
                       % (cond["excel"], cond["value_repr"], extra, br.upper()))
            for c in nd["children"][1:]:
                if c.get("taken"):
                    walk(c)
        elif k == "iferror":
            if nd["detail"].get("caught"):
                out.append("The main expression raised %s, so the fallback is used."
                           % nd["detail"]["caught"])
            for c in nd["children"]:
                if c.get("taken"):
                    walk(c)
        elif k == "logical" and nd["detail"].get("short_circuit_at") is not None:
            out.append("%s short-circuited at argument %d -> %s."
                       % (nd["detail"].get("operator"),
                          nd["detail"]["short_circuit_at"] + 1, nd["value_repr"]))
            for c in nd["children"]:
                walk(c)
        elif k == "lookup" and nd["detail"].get("function") == "VLOOKUP":
            d = nd["detail"]
            if d.get("match_found"):
                out.append("VLOOKUP(%s) matched row %d of %s and returned %s."
                           % (d.get("lookup_repr"), d.get("matched_row_index", 0) + 1,
                              d.get("table_range"), nd["value_repr"]))
            else:
                out.append("VLOOKUP(%s) found no match in %s -> %s."
                           % (d.get("lookup_repr"), d.get("table_range"), nd["value_repr"]))
        else:
            for c in nd.get("children", []):
                walk(c)

    walk(root)
    return out


# =============================================================================
# workbook wrapper
# =============================================================================
class _Book:
    def __init__(self, path=None, data=None):
        if path is None and data is None:
            raise ValueError("workbook path or bytes required")
        self.wb_formula = openpyxl.load_workbook(path or io.BytesIO(data), data_only=False)
        self.wb_values = openpyxl.load_workbook(path or io.BytesIO(data), data_only=True)
        self.sheetnames = list(self.wb_formula.sheetnames)
        self.default_sheet = self.sheetnames[0] if self.sheetnames else None
        self._lc = {s.lower(): s for s in self.sheetnames}

    def real_sheet(self, name):
        if name in self._lc.values():
            return name
        return self._lc.get(str(name).lower(), name)

    def cached(self, sheet, a1):
        try:
            return self.wb_values[self.real_sheet(sheet)][a1].value
        except Exception:  # noqa: BLE001
            return None


# =============================================================================
# the engine
# =============================================================================
class ExcelDeepTraceEngine:
    def __init__(self, workbook_path=None, workbook_bytes=None, filename=None,
                 max_expanded_range=MAX_EXPANDED_RANGE_DEFAULT):
        self.max_expanded_range = int(max_expanded_range)
        self.filename = filename or (os.path.basename(workbook_path) if workbook_path else "workbook.xlsx")
        self._book = _Book(workbook_path, workbook_bytes)
        self.wb_formula = self._book.wb_formula
        self.wb_values = self._book.wb_values
        self.records: Dict[str, Record] = {}
        self._dependents: Dict[str, set] = {}
        self.findings: List[Finding] = []
        self._scanned = False

    # ------------------------------------------------------------------ scan
    def scan(self) -> Dict[str, Any]:
        for ws in self.wb_formula.worksheets:
            sheet = ws.title
            for row in ws.iter_rows():
                for cell in row:
                    v = cell.value
                    if not (isinstance(v, str) and v.startswith("=")):
                        continue
                    a1 = cell.coordinate
                    key = "%s!%s" % (sheet, a1)
                    funcs = _functions_in(v)
                    root, parsed = None, True
                    try:
                        root, _ = parse_excel(v, sheet)
                    except Exception as exc:  # noqa: BLE001
                        parsed = False
                        self.findings.append(Finding(
                            "MEDIUM", "UNPARSED_FORMULA", key,
                            "Could not parse this formula: %s" % exc, {"formula": v}))
                    precs, ranges = _deps(root, sheet) if root is not None else (set(), [])
                    cached = self._book.cached(sheet, a1)
                    rec = Record(
                        key=key, sheet=sheet, cell=a1, coordinate=key, formula=v,
                        cached_value=_json(cached), functions=sorted(funcs),
                        family=_family(funcs, root), formula_length=len(v),
                        dependency_count=len(precs) + len(ranges),
                        is_conditional=bool(funcs & _COND_FUNCS),
                        is_lookup=bool(funcs & _LOOKUP_FUNCS),
                        is_volatile=bool(funcs & _VOLATILE),
                        parsed=parsed, precedents=precs, ranges=ranges, root=root)
                    self.records[key] = rec

        for key, rec in self.records.items():
            for (sh, c1, r1, c2, r2) in rec.ranges:
                rec.precedents |= _range_keys(sh, c1, r1, c2, r2)
            for p in rec.precedents:
                self._dependents.setdefault(p, set()).add(key)

        self._run_findings()
        self._scanned = True
        return self._summary()

    def _summary(self) -> Dict[str, Any]:
        func_usage: Dict[str, int] = {}
        families: Dict[str, int] = {}
        for r in self.records.values():
            families[r.family] = families.get(r.family, 0) + 1
            for fn in r.functions:
                func_usage[fn] = func_usage.get(fn, 0) + 1
        by_sev: Dict[str, int] = {}
        for f in self.findings:
            by_sev[f.severity] = by_sev.get(f.severity, 0) + 1
        sheets = []
        for ws in self.wb_formula.worksheets:
            n = sum(1 for k in self.records if k.startswith(ws.title + "!"))
            sheets.append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column,
                           "formula_cells": n})
        return {
            "engine_build": ENGINE_BUILD,
            "filename": self.filename,
            "sheet_names": list(self.wb_formula.sheetnames),
            "sheets": sheets,
            "total_formula_cells": len(self.records),
            "function_usage": dict(sorted(func_usage.items(), key=lambda kv: -kv[1])),
            "families": families,
            "conditional_cells": sum(1 for r in self.records.values() if r.is_conditional),
            "lookup_cells": sum(1 for r in self.records.values() if r.is_lookup),
            "volatile_cells": sum(1 for r in self.records.values() if r.is_volatile),
            "unparsed_cells": sum(1 for r in self.records.values() if not r.parsed),
            "findings": {"total": len(self.findings), "by_severity": by_sev},
            "max_precedent_depth": self._max_depth_estimate(),
        }

    def _max_depth_estimate(self) -> int:
        memo: Dict[str, int] = {}

        def depth(key, seen):
            if key in memo:
                return memo[key]
            if key in seen or key not in self.records:
                return 0
            d = 1 + max([depth(p, seen | {key}) for p in self.records[key].precedents] or [0])
            memo[key] = d
            return d

        return max([depth(k, set()) for k in self.records] or [0])

    # ---------------------------------------------------------------- findings
    def _run_findings(self) -> None:
        skeletons: Dict[Tuple[str, str], Dict[str, list]] = {}
        for key, rec in self.records.items():
            if rec.is_volatile:
                self.findings.append(Finding(
                    "MEDIUM", "VOLATILE_FUNCTION", key,
                    "Uses a volatile function (%s) - recalculates on every edit."
                    % ", ".join(sorted(set(rec.functions) & _VOLATILE)),
                    {"functions": sorted(set(rec.functions) & _VOLATILE)}))
            if isinstance(rec.cached_value, str) and rec.cached_value in _ERROR_LITERALS:
                self.findings.append(Finding(
                    "HIGH", "CACHED_ERROR", key,
                    "The last saved value of this cell is %s." % rec.cached_value,
                    {"formula": rec.formula, "value": rec.cached_value}))
            if "[" in rec.formula and "]" in rec.formula:
                self.findings.append(Finding(
                    "MEDIUM", "EXTERNAL_REFERENCE", key,
                    "References another workbook - value will be stale outside Excel.",
                    {"formula": rec.formula}))
            if rec.root is not None and "IFERROR" not in rec.functions:
                for nd in rec.root.walk():
                    if isinstance(nd, Bin) and nd.op == "/" and isinstance(nd.right, (CellRef, NameRef)):
                        self.findings.append(Finding(
                            "MEDIUM", "UNGUARDED_DIVISION", key,
                            "Divides by %s without an IFERROR guard." % src(nd.right),
                            {"expression": src(nd)}))
                        break
            if rec.root is not None:
                consts = [float(x.text) for x in rec.root.walk() if isinstance(x, Num)]
                magic = [c for c in consts if c not in (0, 1, -1, 2, 100, 1000, 0.5)]
                has_ref = any(isinstance(x, (CellRef, RangeRef)) for x in rec.root.walk())
                if magic and has_ref and rec.family in ("arithmetic", "conditional", "aggregation"):
                    self.findings.append(Finding(
                        "LOW", "HARDCODED_CONSTANT", key,
                        "Mixes hard-coded number(s) %s with cell references."
                        % ", ".join(str(m) for m in magic[:5]),
                        {"constants": magic[:10], "formula": rec.formula}))
            col_letter = "".join(ch for ch in rec.cell if ch.isalpha())
            skel = _skeletonize(rec.formula)
            grp = skeletons.setdefault((rec.sheet, col_letter), {})
            grp.setdefault(skel, []).append(key)

        for (sheet, col), by_skel in skeletons.items():
            if len(by_skel) < 2:
                continue
            ranked = sorted(by_skel.items(), key=lambda kv: -len(kv[1]))
            majority, majority_keys = ranked[0]
            if len(majority_keys) < 3:
                continue
            for skel, keys in ranked[1:]:
                if len(keys) <= max(1, len(majority_keys) // 5):
                    for k in keys:
                        self.findings.append(Finding(
                            "HIGH", "INCONSISTENT_COLUMN_FORMULA", k,
                            "Formula in column %s of '%s' differs from the %d other rows "
                            "in that column." % (col, sheet, len(majority_keys)),
                            {"this": self.records[k].formula,
                             "typical_pattern": majority}))

        sev_rank = {"HIGH": 0, "MEDIUM": 1, "LOW": 2, "INFO": 3}
        self.findings.sort(key=lambda f: (sev_rank.get(f.severity, 9), f.cell))

    # ---------------------------------------------------------------- keys / values
    def _normalize_key(self, ref: str) -> str:
        ref = str(ref).strip().replace("$", "")
        if "!" in ref:
            sheet, a1 = ref.rsplit("!", 1)
            sheet = _unquote_sheet(sheet)
        else:
            sheet, a1 = self._book.default_sheet, ref
        return "%s!%s" % (self._book.real_sheet(sheet), a1.upper())

    def value(self, key: str) -> Any:
        k = self._normalize_key(key)
        sheet, a1 = k.split("!", 1)
        return _json(self._book.cached(sheet, a1))

    def _defined_name(self, name: str):
        return None  # named ranges are resolved lazily elsewhere; kept simple for now

    def _grid(self, rangeref: RangeRef, cap: int) -> List[list]:
        sheet = rangeref.sheet or self._book.default_sheet
        ws = self._book.wb_values[self._book.real_sheet(sheet)]
        lo_c, hi_c = sorted((col_to_idx(rangeref.c1), col_to_idx(rangeref.c2)))
        if rangeref.whole_column:
            lo_r, hi_r = 1, max(ws.max_row or 1, 1)
        else:
            lo_r, hi_r = sorted((rangeref.r1, rangeref.r2))
        if (hi_c - lo_c + 1) * (hi_r - lo_r + 1) > cap:
            raise ValueError("range %s exceeds max_expanded_range (%d)" % (rangeref.a1, cap))
        return [[ws.cell(row=r, column=c).value for c in range(lo_c, hi_c + 1)]
                for r in range(lo_r, hi_r + 1)]

    def _grid_or_err(self, rangeref: RangeRef):
        try:
            return self._grid(rangeref, self.max_expanded_range)
        except Exception:  # noqa: BLE001
            return XlError("#REF!")

    # ---------------------------------------------------------------- graph trace
    def trace(self, key: str, direction: str = "precedents",
              max_depth: int = 12, max_nodes: int = 500) -> Dict[str, Any]:
        key = self._normalize_key(key)
        nodes: Dict[str, dict] = {}
        edges: List[dict] = []
        seen = {key}
        frontier = [(key, 0)]
        truncated = False
        while frontier:
            cur, d = frontier.pop(0)
            nodes[cur] = self._node_meta(cur, d)
            if d >= max_depth:
                truncated = True
                continue
            if direction == "precedents":
                nbrs = self.records[cur].precedents if cur in self.records else set()
            else:
                nbrs = self._dependents.get(cur, set())
            for nb in sorted(nbrs):
                if direction == "precedents":
                    edges.append({"from": nb, "to": cur})
                else:
                    edges.append({"from": cur, "to": nb})
                if nb not in seen:
                    if len(seen) >= max_nodes:
                        truncated = True
                        continue
                    seen.add(nb)
                    frontier.append((nb, d + 1))
        return {"root": key, "direction": direction, "nodes": list(nodes.values()),
                "edges": edges, "truncated": truncated, "node_count": len(nodes),
                "max_depth": max(n["depth"] for n in nodes.values()) if nodes else 0}

    def _node_meta(self, key: str, depth: int) -> dict:
        sheet, a1 = key.split("!", 1)
        rec = self.records.get(key)
        cached = self._book.cached(sheet, a1)
        return {
            "id": key, "sheet": sheet, "cell": a1, "depth": depth,
            "value": _json(cached), "value_repr": _repr(cached),
            "is_formula": rec is not None,
            "formula": rec.formula if rec else None,
            "family": rec.family if rec else "input",
            "functions": rec.functions if rec else [],
            "is_conditional": rec.is_conditional if rec else False,
            "is_lookup": rec.is_lookup if rec else False,
        }

    # ---------------------------------------------------------------- analyses
    def _root_or_none(self, key: str) -> Tuple[str, Optional[Record]]:
        k = self._normalize_key(key)
        return k, self.records.get(k)

    def branch_analysis(self, key: str) -> Optional[Dict[str, Any]]:
        k, rec = self._root_or_none(key)
        if rec is None or rec.root is None:
            return None
        ifs = [n for n in rec.root.walk() if isinstance(n, If)]
        if not ifs:
            return {"has_branches": False, "branches": []}
        ev = _Eval(self, trace=False)
        out = []
        for i, n in enumerate(ifs):
            try:
                cv = ev.value(n.cond)
                taken = _truthy(cv)
                out.append({
                    "index": i,
                    "condition": src(n.cond),
                    "condition_value": _json(cv),
                    "condition_repr": (cv.code if _is_err(cv) else _repr(cv)),
                    "branch_taken": "ERROR" if _is_err(cv) else ("TRUE" if taken else "FALSE"),
                    "if_true": src(n.then),
                    "if_false": src(n.els) if n.els is not None else "FALSE",
                })
            except Exception as exc:  # noqa: BLE001
                out.append({"index": i, "condition": src(n.cond), "error": str(exc)})
        return {"has_branches": True, "branches": out}

    def lookup_analysis(self, key: str) -> Optional[Dict[str, Any]]:
        k, rec = self._root_or_none(key)
        if rec is None or rec.root is None:
            return None
        calls = [n for n in rec.root.walk()
                 if isinstance(n, Call) and n.name.upper() in _LOOKUP_FUNCS]
        if not calls:
            return {"has_lookups": False, "lookups": []}
        ev = _Eval(self, trace=True)
        out = []
        for n in calls:
            try:
                node, _ = ev.ev(n, 0, "root")
                out.append(node.get("detail", {}) or {"function": n.name.upper()})
            except Exception as exc:  # noqa: BLE001
                out.append({"function": n.name.upper(), "error": str(exc)})
        return {"has_lookups": True, "lookups": out}

    def expression_tree(self, key: str, max_depth: int = MAX_TREE_DEPTH_DEFAULT) -> Dict[str, Any]:
        k = self._normalize_key(key)
        sheet, a1 = k.split("!", 1)
        rec = self.records.get(k)
        cached = self._book.cached(sheet, a1)
        if rec is None:
            return {"key": k, "is_formula": False, "value": _json(cached),
                    "value_repr": _repr(cached),
                    "header": self._header_for(sheet, col_to_idx(a1.rstrip("0123456789")))}
        if rec.root is None:
            return {"key": k, "is_formula": True, "parse_error": True,
                    "formula": rec.formula, "value": _json(cached), "value_repr": _repr(cached)}
        if_ids = {id(n): i for i, n in enumerate(x for x in rec.root.walk() if isinstance(x, If))}
        ev = _Eval(self, trace=True, max_depth=max_depth, if_ids=if_ids)
        root, val = ev.ev(rec.root, 0, "root")
        no_excel_value = _is_empty(cached)
        if no_excel_value:
            # workbook stored no cached result for this cell - nothing to check against
            mismatch = False
        else:
            mismatch = not _values_close(val.code if _is_err(val) else val, cached)
            if _is_err(val) and isinstance(cached, str) and cached in _ERROR_LITERALS:
                mismatch = val.code != cached
            elif _is_err(val):
                mismatch = True
        return {
            "key": k, "is_formula": True, "formula": rec.formula,
            "family": rec.family, "functions": rec.functions,
            "value": _json(val), "value_repr": (val.code if _is_err(val) else _repr(val)),
            "excel_value": _json(cached), "excel_value_repr": _repr(cached),
            "tracer_reproduced_excel": not mismatch,
            "no_excel_value": no_excel_value,
            "notes": ev.notes,
            "narrative": _narrate(root, k, cached),
            "root": root,
        }

    def row_journey(self, sheet: str, row: int) -> Dict[str, Any]:
        if sheet not in self.wb_formula.sheetnames:
            raise ValueError("worksheet %r does not exist" % sheet)
        if not self._scanned:
            self.scan()
        ws = self.wb_formula[sheet]
        row = int(row)
        cells, inputs = [], []
        for c in range(1, (ws.max_column or 1) + 1):
            a1 = "%s%d" % (idx_to_col(c), row)
            key = "%s!%s" % (sheet, a1)
            rec = self.records.get(key)
            cached = self._book.cached(sheet, a1)
            header = self._header_for(sheet, c)
            if rec is None:
                if not _is_empty(cached):
                    inputs.append({"cell": a1, "header": header,
                                   "value": _json(cached), "value_repr": _repr(cached)})
                continue
            ba = self.branch_analysis(key) or {}
            la = self.lookup_analysis(key) or {}
            tree = self.expression_tree(key)
            shown = cached if not _is_empty(cached) else tree.get("value")
            cells.append({
                "cell": a1, "key": key, "header": header,
                "formula": rec.formula, "value": _json(shown), "value_repr": _repr(shown),
                "family": rec.family, "functions": rec.functions,
                "is_conditional": rec.is_conditional, "is_lookup": rec.is_lookup,
                "branches": ba.get("branches", []),
                "lookups": la.get("lookups", []),
                "precedents": sorted(rec.precedents)[:40],
                "tracer_reproduced_excel": tree.get("tracer_reproduced_excel", True),
                "narrative": tree.get("narrative", []),
            })
        return {"sheet": sheet, "row": row, "formula_cells": len(cells),
                "input_cells": len(inputs), "cells": cells, "inputs": inputs}

    def compare_cells(self, left: str, right: str) -> Dict[str, Any]:
        lk, lrec = self._root_or_none(left)
        rk, rrec = self._root_or_none(right)
        lv = self._cell_value_or_eval(lk)
        rv = self._cell_value_or_eval(rk)
        ln, rn = _num(lv), _num(rv)
        diff = (float(ln) - float(rn)) if isinstance(ln, (int, float)) and isinstance(rn, (int, float)) else None
        lp = lrec.precedents if lrec else set()
        rp = rrec.precedents if rrec else set()
        return {
            "left": {"key": lk, "formula": lrec.formula if lrec else None,
                     "value": _json(lv), "value_repr": _repr(lv),
                     "functions": lrec.functions if lrec else [],
                     "family": lrec.family if lrec else "input"},
            "right": {"key": rk, "formula": rrec.formula if rrec else None,
                      "value": _json(rv), "value_repr": _repr(rv),
                      "functions": rrec.functions if rrec else [],
                      "family": rrec.family if rrec else "input"},
            "values_match": _values_close(lv, rv),
            "difference": diff,
            "formula_identical": bool(lrec and rrec
                                      and _skeletonize(lrec.formula) == _skeletonize(rrec.formula)),
            "shared_precedents": sorted(lp & rp),
            "only_left_precedents": sorted(lp - rp),
            "only_right_precedents": sorted(rp - lp),
        }

    def narrate(self, key: str) -> List[str]:
        return self.expression_tree(key).get("narrative", [])

    # ---------------------------------------------------------------- evaluated values
    def _cell_value_or_eval(self, key: str) -> Any:
        """The workbook's cached value if it has one, otherwise evaluate the cell's
        formula. This is what makes formula-only workbooks (no cached results, which is
        what openpyxl writes) usable for comparison and the row journey."""
        sheet, a1 = key.split("!", 1)
        cached = self._book.cached(sheet, a1)
        if not _is_empty(cached) or key not in self.records:
            return cached
        rec = self.records[key]
        if rec.root is None:
            return cached
        try:
            v = _Eval(self, trace=False, max_depth=64).value(rec.root)
        except Exception:  # noqa: BLE001
            return cached
        return cached if _is_err(v) else v

    def evaluate_sheet(self, sheet: str, max_rows: int = 20000) -> Dict[str, Any]:
        """The sheet as a table with **every formula evaluated** (input cells kept as-is).
        Returns ``{sheet, header_row, headers, rows}`` - the backend turns it into a
        DataFrame for the comparison."""
        info = self.sheet_columns(sheet)
        real, hr = info["sheet"], info["header_row"]
        ws = self._book.wb_values[real]
        cols = info["columns"]
        headers = [c["name"] or ("Column_%s" % c["letter"]) for c in cols]
        last = min(ws.max_row or hr, hr + int(max_rows))
        rows: List[list] = []
        for r in range(hr + 1, last + 1):
            vals, blank = [], True
            for c in cols:
                key = "%s!%s%d" % (real, c["letter"], r)
                if key in self.records:
                    v = self._cell_value_or_eval(key)
                else:
                    v = ws.cell(row=r, column=c["index"]).value
                if not _is_empty(v):
                    blank = False
                vals.append(_json(v))
            if not blank:
                rows.append(vals)
        return {"sheet": real, "header_row": hr, "headers": headers, "rows": rows}

    # ---------------------------------------------------------------- columns / mismatch
    def sheet_columns(self, sheet: str) -> Dict[str, Any]:
        real = self._book.real_sheet(sheet)
        ws = self._book.wb_values[real]
        ncol = ws.max_column or 1
        header_row = 1
        for r in range(1, min(ws.max_row or 1, 8) + 1):
            txt = [ws.cell(row=r, column=c).value for c in range(1, ncol + 1)]
            if sum(1 for v in txt if isinstance(v, str) and v.strip()) >= 2:
                header_row = r
                break
        cols = []
        for c in range(1, ncol + 1):
            v = ws.cell(row=header_row, column=c).value
            cols.append({"index": c, "letter": idx_to_col(c),
                         "name": str(v).strip() if v is not None else ""})
        return {"sheet": real, "header_row": header_row, "columns": cols}

    def mismatch_rows(self, sheet: str, left: str, right: str,
                      tolerance: float = 0.01, limit: int = 1000) -> Dict[str, Any]:
        info = self.sheet_columns(sheet)

        def resolve(spec):
            spec = str(spec).strip()
            for c in info["columns"]:
                if c["letter"] == spec.upper() or c["name"].lower() == spec.lower():
                    return c["index"]
            raise ValueError("column %r not found on sheet %r" % (spec, info["sheet"]))

        lc, rc = resolve(left), resolve(right)
        real = info["sheet"]
        ws = self._book.wb_values[real]
        ll, rl = idx_to_col(lc), idx_to_col(rc)
        out = []
        for r in range(info["header_row"] + 1, (ws.max_row or 1) + 1):
            lv = self._cell_value_or_eval("%s!%s%d" % (real, ll, r))
            rv = self._cell_value_or_eval("%s!%s%d" % (real, rl, r))
            if _is_empty(lv) and _is_empty(rv):
                continue
            ln, rn = _num(lv), _num(rv)
            if isinstance(ln, (int, float)) and isinstance(rn, (int, float)):
                d = float(ln) - float(rn)
                if abs(d) > float(tolerance):
                    out.append({"row": r, "left": _json(lv), "right": _json(rv), "diff": d})
            elif str("" if lv is None else lv).strip() != str("" if rv is None else rv).strip():
                out.append({"row": r, "left": _json(lv), "right": _json(rv), "diff": None})
            if len(out) >= limit:
                break
        return {"sheet": info["sheet"], "left": left, "right": right,
                "tolerance": float(tolerance), "rows": out, "total": len(out)}

    # ---------------------------------------------------------------- misc
    def _header_for(self, sheet: str, col_idx: int) -> Optional[str]:
        try:
            for r in (1, 2, 3):
                v = self._book.wb_values[sheet].cell(row=r, column=col_idx).value
                if isinstance(v, str) and v.strip():
                    return v.strip()
        except Exception:  # noqa: BLE001
            pass
        return None


_SKEL_NUM = re.compile(r"\d+")
_SKEL_ROW = re.compile(r"(\$?[A-Za-z]{1,3}\$?)\d+")


def _skeletonize(formula: str) -> str:
    """Normalise a formula so sibling rows compare equal: strip row numbers + literals."""
    s = _SKEL_ROW.sub(r"\1R", formula)
    s = re.sub(r'"[^"]*"', '"S"', s)
    s = _SKEL_NUM.sub("N", s)
    return re.sub(r"\s+", "", s).upper()
