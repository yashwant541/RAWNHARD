"""Analyse an uploaded sample workbook.

The sample ``.xlsx`` holds, on separate sheets:
  * one **data sheet** whose rows may start at any row (header auto-detected);
  * some columns of that data sheet contain **Excel formulas** -> "computed" columns;
  * other sheets are **lookup / mapping tables** referenced by the formulas;
  * formulas may reference **toggle** names that are neither columns nor lookup sheets.

``analyze_sample(bytes)`` returns a plain dict the Admin UI renders for review/editing.
Only openpyxl + (optional) the sibling ``formula_translate`` module are used - no dataiku.
"""

from __future__ import annotations

import io
import re
import string
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

from . import formula_translate

MAX_SCAN_ROWS = 40
_CELL_RE = re.compile(r"(\$?)([A-Za-z]{1,3})(\$?)(\d+)")
_SHEET_REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!")
# A sheet-qualified single cell OR range (e.g. "Rates!$A$2" / "Rates!$A$2:$B$4") - matched
# as one unit so range-generalisation never mistakes the range's second cell for a
# same-sheet "fixed cell" (see generalize_formula).
_SHEET_QUALIFIED_RE = re.compile(
    _SHEET_REF_RE.pattern + _CELL_RE.pattern + r"(?::" + _CELL_RE.pattern + r")?"
)
_FUNC_RE = re.compile(r"([A-Za-z][A-Za-z0-9_.]*)\s*\(")
_TOKEN_RE = re.compile(r"[A-Za-z_][A-Za-z0-9_]*")
# sheets named "Toggle", "Toggle 1", "Toggles 2" ... hold the Yes/No switches (like the
# lookup/mapping sheets hold reference tables).
_TOGGLE_SHEET_RE = re.compile(r"^\s*toggles?\s*[-_ ]?\d*\s*$", re.I)
_YESNO = {"yes": "Yes", "y": "Yes", "true": "Yes", "1": "Yes",
          "no": "No", "n": "No", "false": "No", "0": "No"}
_GENERIC_LABELS = {"name", "value", "toggle", "toggles", "key", "parameter",
                   "param", "setting", "switch", "flag"}


def is_toggle_sheet(name: str) -> bool:
    return bool(_TOGGLE_SHEET_RE.match(name or ""))


def _split_sheet_cell(text: str) -> Tuple[str, str]:
    sheet, a1 = text.split("!", 1)
    if sheet.startswith("'") and sheet.endswith("'"):
        sheet = sheet[1:-1]
    return sheet, a1.replace("$", "")

# Excel built-ins we translate / recognise; never treated as toggles.
KNOWN_FUNCS = {
    "IF", "IFS", "IFERROR", "AND", "OR", "NOT", "ROUND", "ROUNDUP", "ROUNDDOWN",
    "ABS", "MIN", "MAX", "SUM", "AVERAGE", "COUNT", "INT", "MOD", "SQRT", "POWER",
    "VLOOKUP", "HLOOKUP", "INDEX", "MATCH", "LOOKUP", "XLOOKUP",
    "CONCAT", "CONCATENATE", "LEFT", "RIGHT", "MID", "LEN", "UPPER", "LOWER", "TRIM",
    "TEXT", "VALUE", "TODAY", "NOW", "YEAR", "MONTH", "DAY", "DATE", "TRUE", "FALSE",
    "ISBLANK", "ISNUMBER", "ISERROR", "COALESCE",
}


# --------------------------------------------------------------------------- utils
def col_letter_to_index(letter: str) -> int:
    """``A`` -> 0, ``B`` -> 1, ``AA`` -> 26 ..."""
    n = 0
    for ch in letter.upper():
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def col_index_to_letter(idx: int) -> str:
    idx += 1
    out = ""
    while idx > 0:
        idx, rem = divmod(idx - 1, 26)
        out = string.ascii_uppercase[rem] + out
    return out


def _norm(v: Any) -> str:
    return re.sub(r"\s+", " ", str(v)).strip() if v is not None else ""


def _looks_like_label(v: Any) -> bool:
    s = _norm(v)
    if not s:
        return False
    try:
        float(s.replace(",", ""))
        return False
    except ValueError:
        return True


def detect_header_row(rows: List[List[Any]], max_scan: int = MAX_SCAN_ROWS) -> int:
    """Return the 0-based index of the most likely header row.

    Heuristic: the row with the most distinct non-empty text cells that also has a
    non-empty data row beneath it and no empty cells between its populated columns.
    """
    best_idx, best_score = 0, -1.0
    limit = min(len(rows), max_scan)
    for i in range(limit):
        row = rows[i]
        labels = [c for c in row if _looks_like_label(c)]
        non_null = [c for c in row if _norm(c)]
        if len(labels) < 2:
            continue
        nxt = rows[i + 1] if i + 1 < len(rows) else []
        has_data_below = any(_norm(c) for c in nxt)
        distinct = len({_norm(c).lower() for c in labels})
        score = distinct + 0.5 * len(non_null)
        if has_data_below:
            score += 3
        if len(labels) == len(non_null):  # header row is usually all-text
            score += 1
        if score > best_score:
            best_idx, best_score = i, score
    return best_idx


def _sheet_rows(ws, values_only: bool = True) -> List[List[Any]]:
    return [list(r) for r in ws.iter_rows(values_only=values_only)]


def _headers_from_row(row: List[Any]) -> List[str]:
    headers, seen = [], {}
    for j, cell in enumerate(row):
        name = _norm(cell) or f"Column{col_index_to_letter(j)}"
        if name in seen:
            seen[name] += 1
            name = f"{name}.{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)
    return headers


# ------------------------------------------------------------------- formula tools
_TOK_PLACEHOLDER_RE = re.compile(r"\{tok:[^}]*\}")


def extract_refs(formula: str) -> Dict[str, List[str]]:
    """Split a formula into referenced sheet names, function names and bare tokens."""
    sheets = sorted({m.group(1) or m.group(2) for m in _SHEET_REF_RE.finditer(formula)})
    funcs = sorted({m.group(1).upper() for m in _FUNC_RE.finditer(formula)})
    # Strip string literals, sheet-qualified refs and {tok:...} placeholders before
    # scanning bare tokens (a placeholder's own text must not look like a new token).
    stripped = re.sub(r'"[^"]*"', "", formula)
    stripped = _SHEET_REF_RE.sub("", stripped)
    stripped = _TOK_PLACEHOLDER_RE.sub("", stripped)
    stripped = _CELL_RE.sub("", stripped)
    tokens = sorted(
        {
            t for t in _TOKEN_RE.findall(stripped)
            if t.upper() not in KNOWN_FUNCS and not re.fullmatch(r"[A-Za-z]{1,3}\d*", t)
        }
    )
    return {"sheets": sheets, "funcs": funcs, "tokens": tokens}


def generalize_formula(
    formula: str, row: int, header_letters: Dict[str, str],
    fixed_cell_hook: Optional[Any] = None,
    sheet_cell_hook: Optional[Any] = None,
) -> str:
    """Replace same-sheet cell refs on ``row`` with the ``{r}`` placeholder.

    ``A5`` on row 5 -> ``A{r}``. Any other same-sheet cell ref - an absolute ``$A$5``, or
    a plain ref that points at a different row (a stray look-elsewhere reference) - is
    "fixed": it stays constant across rows, so it can't become ``{r}``. Such a ref is
    typically a single labelled parameter/toggle cell placed elsewhere on the sheet
    (e.g. ``IF($AB$2="Yes", ...)``). By default it is left as-is; pass ``fixed_cell_hook``
    (called with ``(raw_text, col_letters, row_num)``) to substitute something else, e.g.
    a ``{tok:Name}`` placeholder once that cell has been registered as a toggle.

    ``sheet_cell_hook(sheet_name, cell_a1, raw_text)`` is called for every *sheet-qualified
    single cell* (``'Toggle 1'!$B$2``); return a replacement string (e.g. ``{tok:Name}``)
    or ``None`` to leave it untouched. Sheet-qualified *ranges* (VLOOKUP tables) are always
    left alone.
    """
    def repl(m: re.Match) -> str:
        dollar_col, col, dollar_row, rownum = m.groups()
        if dollar_row != "$" and int(rownum) == row:
            return f"{dollar_col}{col}{{r}}"
        if fixed_cell_hook:
            return fixed_cell_hook(m.group(0), col.upper(), int(rownum))
        return m.group(0)

    # Protect sheet-qualified refs from row generalisation (but let the sheet-cell hook
    # rewrite single-cell refs, e.g. a toggle-sheet reference -> {tok:Name}).
    placeholders: Dict[str, str] = {}

    def stash(m: re.Match) -> str:
        text = m.group(0)
        if sheet_cell_hook and ":" not in text:
            sheet, a1 = _split_sheet_cell(text)
            replacement = sheet_cell_hook(sheet, a1, text)
            if replacement is not None:
                return replacement
        key = f"\x00{len(placeholders)}\x00"
        placeholders[key] = text
        return key

    protected = _SHEET_QUALIFIED_RE.sub(stash, formula)
    generalized = _CELL_RE.sub(repl, protected)
    for key, val in placeholders.items():
        generalized = generalized.replace(key, val)
    return generalized


def _guess_cell_label(ws_v, row_num: int, col_idx: int) -> Optional[str]:
    """A fixed cell is usually a value next to a text label - e.g. 'Include Tax' in the
    cell to its left, or above it. Used to name the toggle after that label."""
    candidates = []
    if col_idx > 0:
        candidates.append(ws_v.cell(row=row_num, column=col_idx).value)  # left neighbour
    if row_num > 1:
        candidates.append(ws_v.cell(row=row_num - 1, column=col_idx + 1).value)  # above
    for v in candidates:
        if _looks_like_label(v):
            return _norm(v)
    return None


def _guess_toggle_value(raw: Any) -> str:
    s = _norm(raw).lower()
    return "Yes" if s in ("yes", "true", "y", "1") else "No"


# ------------------------------------------------------------------------- analyse
def analyze_sample(file_bytes: bytes, data_sheet: Optional[str] = None) -> Dict[str, Any]:
    """Parse the sample workbook. See module docstring for the return shape."""
    wb_f = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    wb_v = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    warnings: List[str] = []

    # 1. pick the data sheet: explicit choice, else the sheet with the most formula cells,
    #    else the sheet with the largest populated area.
    def formula_count(ws) -> int:
        c = 0
        for r in ws.iter_rows():
            for cell in r:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    c += 1
        return c

    sheet_names = wb_f.sheetnames
    if data_sheet and data_sheet in sheet_names:
        ds_name = data_sheet
    else:
        scored = sorted(
            sheet_names,
            key=lambda n: (formula_count(wb_f[n]), wb_f[n].max_row * wb_f[n].max_column),
            reverse=True,
        )
        ds_name = scored[0]
        if formula_count(wb_f[ds_name]) == 0:
            warnings.append(
                "No Excel formula cells were found on any sheet; "
                f"'{ds_name}' was chosen by size. Pick the data sheet manually if wrong."
            )

    ws_f, ws_v = wb_f[ds_name], wb_v[ds_name]
    frows = _sheet_rows(ws_f)
    vrows = _sheet_rows(ws_v)
    if not frows:
        raise ValueError(f"Data sheet '{ds_name}' is empty.")

    header_row = detect_header_row(vrows)
    first_data = header_row + 1
    if first_data >= len(frows):
        raise ValueError(f"Data sheet '{ds_name}' has a header but no data rows.")

    # Keep only columns that are actually part of the table: a header label, or data
    # underneath it. openpyxl pads every row out to the sheet's overall max_column, so
    # without this a cell placed far to the right (e.g. a toggle at $AB$2) would drag in
    # dozens of blank "ColumnF".."ColumnAB" placeholders as bogus canonical columns.
    raw_header_row = frows[header_row]

    def _col_has_data(j: int) -> bool:
        return any(j < len(r) and _norm(r[j]) for r in frows[first_data:])

    kept = [j for j in range(len(raw_header_row))
           if _norm(raw_header_row[j]) or _col_has_data(j)]
    headers = _headers_from_row([raw_header_row[j] for j in kept])
    letter_by_header = {h: col_index_to_letter(kept[i]) for i, h in enumerate(headers)}
    header_by_letter = {col_index_to_letter(kept[i]): h for i, h in enumerate(headers)}

    # 2. toggles + column classification.
    #
    # Toggles (Yes/No switches used in formulas) can come from:
    #   (a) a dedicated "Toggle" / "Toggle 1" ... sheet holding name/value rows, OR
    #   (b) a single labelled cell placed elsewhere on the data sheet
    #       (``IF($AB$2="Yes", ...)``), OR
    #   (c) a bare word used directly in a formula (``IF(IncludeTax, ...)``).
    # All three feed one ``toggle_defs`` map: name -> {name, value, [sheet], [cell]}.
    # Formula references to (a)/(b) are templated to ``{tok:Name}`` so they survive both
    # the pandas translation and rebuilding the live Excel formula.
    toggle_defs: Dict[str, Dict[str, Any]] = {}
    used_names: Dict[str, str] = {}
    sheet_cell_to_toggle: Dict[Tuple[str, Optional[str]], str] = {}
    toggle_sheet_names = {
        n for n in sheet_names if n != ds_name and is_toggle_sheet(n)
    }

    def _register_toggle(name: str, value: str, sheet=None, cell=None, tag=None) -> str:
        tag = tag or (f"{sheet}!{cell}" if sheet else cell or name)
        base, cand = name, name
        i = 2
        while used_names.get(cand, tag) != tag:
            cand = f"{base} ({i})"
            i += 1
        used_names[cand] = tag
        entry = {"name": cand, "value": value}
        if sheet:
            entry["sheet"] = sheet
        if cell:
            entry["cell"] = cell
        toggle_defs.setdefault(cand, entry)
        return cand

    for tname in sorted(toggle_sheet_names):
        trows = _sheet_rows(wb_v[tname])
        got = False
        for ri, trow in enumerate(trows):
            key = _norm(trow[0]) if trow else ""
            val = _norm(trow[1]) if len(trow) > 1 else ""
            if key and key.lower() not in _GENERIC_LABELS and val.lower() in _YESNO:
                nm = _register_toggle(key, _YESNO[val.lower()], sheet=tname,
                                      cell=f"B{ri + 1}")
                sheet_cell_to_toggle[(tname, f"B{ri + 1}")] = nm
                sheet_cell_to_toggle.setdefault((tname, None), nm)
                got = True
        if not got:  # fall back: first Yes/No cell anywhere on the sheet
            for ri, trow in enumerate(trows):
                for ci, cell in enumerate(trow):
                    if _norm(cell).lower() in _YESNO:
                        addr = f"{col_index_to_letter(ci)}{ri + 1}"
                        nm = _register_toggle(tname, _YESNO[_norm(cell).lower()],
                                              sheet=tname, cell=addr)
                        sheet_cell_to_toggle[(tname, addr)] = nm
                        sheet_cell_to_toggle.setdefault((tname, None), nm)
                        got = True
                        break
                if got:
                    break

    def _fixed_cell_hook(raw_text: str, col_letters: str, row_num: int) -> str:
        addr = f"{col_letters}{row_num}"
        key = ("__data__", addr)
        if key not in sheet_cell_to_toggle:
            col_idx = col_letter_to_index(col_letters)
            try:
                raw_value = ws_v.cell(row=row_num, column=col_idx + 1).value
            except Exception:  # noqa: BLE001
                raw_value = None
            label = _guess_cell_label(ws_v, row_num, col_idx) or f"Cell {addr}"
            sheet_cell_to_toggle[key] = _register_toggle(
                label, _guess_toggle_value(raw_value), cell=addr, tag=addr
            )
        return "{tok:%s}" % sheet_cell_to_toggle[key]

    def _sheet_cell_hook(sheet: str, a1: str, raw_text: str):
        if sheet not in toggle_sheet_names:
            return None  # a lookup-sheet cell, or anything else - leave it alone
        name = (sheet_cell_to_toggle.get((sheet, a1))
                or sheet_cell_to_toggle.get((sheet, None)))
        if not name:
            name = _register_toggle(sheet, "Yes", sheet=sheet, cell=a1)
            sheet_cell_to_toggle[(sheet, a1)] = name
        return "{tok:%s}" % name

    canonical: List[str] = []
    computed: List[Dict[str, Any]] = []
    for j, name in enumerate(headers):
        orig_j = kept[j]
        formula_cell = None
        formula_excel_row = None
        for ridx in range(first_data, len(frows)):
            val = frows[ridx][orig_j] if orig_j < len(frows[ridx]) else None
            if isinstance(val, str) and val.startswith("="):
                formula_cell = val
                formula_excel_row = ridx + 1  # openpyxl rows are 1-based
                break
        if formula_cell is None:
            canonical.append(name)
            continue
        generalized = generalize_formula(
            formula_cell, formula_excel_row, letter_by_header,
            _fixed_cell_hook, _sheet_cell_hook,
        )
        pandas_expr, xl_notes = formula_translate.translate(
            generalized, header_by_letter
        )
        refs = extract_refs(generalized)
        computed.append(
            {
                "name": name,
                "position": j,
                "excel_formula": generalized,
                "sample_excel_formula": formula_cell,
                "pandas_expr": pandas_expr,
                "refs": refs,
                "notes": xl_notes,
            }
        )

    if not computed:
        warnings.append("No computed (formula) columns detected on the data sheet.")

    # 3. lookup sheets = every other sheet that isn't a toggle sheet
    lookups: List[Dict[str, Any]] = []
    for name in sheet_names:
        if name == ds_name or name in toggle_sheet_names:
            continue
        lrows = _sheet_rows(wb_v[name])
        if not any(any(_norm(c) for c in row) for row in lrows):
            continue
        lhr = detect_header_row(lrows)
        lheaders = _headers_from_row(lrows[lhr]) if lrows else []
        preview = [
            {lheaders[k]: row[k] if k < len(row) else None for k in range(len(lheaders))}
            for row in lrows[lhr + 1: lhr + 6]
        ]
        lookups.append(
            {
                "sheet": name,
                "header_row": lhr,
                "columns": lheaders,
                "preview": preview,
                "n_rows": max(0, len(lrows) - lhr - 1),
            }
        )
    lookup_sheet_names = {l["sheet"] for l in lookups}

    # 4. bare-word tokens -> toggle candidates, merged with the sheet/fixed-cell toggles
    # already found while generalizing the formulas (step 2, above).
    canon_lower = {c.lower() for c in canonical}
    comp_lower = {c["name"].lower() for c in computed}
    for c in computed:
        for tok in c["refs"]["tokens"]:
            tl = tok.lower()
            if tl in canon_lower or tl in comp_lower:
                continue
            if tok in lookup_sheet_names or tok in toggle_sheet_names:
                continue
            if tok not in toggle_defs:
                toggle_defs[tok] = {"name": tok, "value": "Yes"}
        for sh in c["refs"]["sheets"]:
            if sh not in lookup_sheet_names and sh not in toggle_sheet_names:
                warnings.append(
                    f"Column '{c['name']}' references sheet '{sh}' which is not a "
                    "lookup sheet in the workbook."
                )

    # 5. name-collision guard (computed columns must be distinct from data columns)
    collisions = comp_lower & canon_lower
    if collisions:
        warnings.append(
            "Computed columns reuse data-column names: "
            + ", ".join(sorted(collisions))
            + ". Rename them so comparisons can address each side."
        )

    return {
        "sheets": sheet_names,
        "data_sheet": ds_name,
        "header_row": header_row,
        "canonical_schema": canonical,
        "computed_columns": computed,
        "lookups": lookups,
        "toggle_sheets": sorted(toggle_sheet_names),
        "toggles": list(toggle_defs.values()),
        "warnings": warnings,
    }


def sheet_to_records(file_bytes: bytes, sheet: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    """Return ``(headers, rows-as-dicts)`` for one sheet, header auto-detected.

    Used by the backend to persist each lookup sheet into the mapping folder.
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    rows = _sheet_rows(wb[sheet])
    hr = detect_header_row(rows)
    headers = _headers_from_row(rows[hr]) if rows else []
    records = []
    for row in rows[hr + 1:]:
        if not any(_norm(c) for c in row):
            continue
        records.append(
            {headers[k]: (row[k] if k < len(row) else None) for k in range(len(headers))}
        )
    return headers, records
